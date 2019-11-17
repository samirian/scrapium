from multiprocessing import Process
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import csv
from openpyxl import Workbook
from samir.scrapium.chrome_options import Chrome_Options


class Scrapium(Process):
	def __init__(self, driver_name: str):
		"""Creates a process that includes a driver.

		driver_name = 'chrome' -> chromedriver
		"""
		self.state_fieldnames = []
		self.state_filename = None
		Process.__init__(self)
		if driver_name == 'chrome':
			self.options = Chrome_Options()

	def open_url(self, url: str, trials: int = 1) -> bool:
		"""Tries to open the url page for trials times.
		
		Returns True if the page is loaded successfully
		and False othewise.
		"""
		for i in range(trials):
			try:
				self.driver.get(url)
				return True
			except TimeoutException:
				pass
		return False

	def new_tab(self):
		self.driver.execute_script('window.open("","_blank");')

	def jump_to_tap(self, tab_index: int):
		self.driver.switch_to.window(self.driver.window_handles[tab_index])

	def wait_for_element_visibility(self, xpath: str):
		self.wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))

	def click_element_by_xpath(self, xpath: str):
		self.wait_for_element_visibility(xpath)
		self.driver.find_element_by_xpath(xpath).click()

	def click_element_by_xpath_using_actions(self, xpath: str):
		"""Move to element`s postion, then click it using actions.
		"""
		self.wait_for_element_visibility(xpath)
		element = self.driver.find_element_by_xpath(xpath)
		self.actions.move_to_element(element).click().perform()

	def get_text_from_element_by_xpath(self, xpath: str) -> str:
		"""Get text from the element specified by xpath.
		"""
		self.wait_for_element_visibility(xpath)
		return self.driver.find_element_by_xpath(xpath).text

	def get_attribute_from_element_by_xpath(self, attribute_name: str, xpath: str) -> str:
		"""Get an attribute value from element specified by xpath.
		"""
		self.wait_for_element_visibility(xpath)
		return self.driver.find_element_by_xpath(xpath).get_attribute(attribute_name)

	def get_attribute_from_element(self, attribute_name: str, element) -> str:
		return element.get_attribute(attribute_name)
		
	def scroll_to_end_of_page(self):
		self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

	def scroll_to_top_of_page(self):
		self.driver.execute_script("window.scrollTo(0, 0);")

	def scroll_to_view(self, xpath: str=None, element=None, offset=0):
		if xpath != None:
			self.wait_for_element_visibility(xpath)
			element = self.driver.find_element_by_xpath(xpath)
		self.driver.execute_script(
			"""
			//arguments[0].scrollIntoView(true);
			var bodyRect = document.body.getBoundingClientRect(),
			elemRect = arguments[0].getBoundingClientRect(),
			offset   = elemRect.top - bodyRect.top - arguments[1];
			window.scrollTo(0, offset);
			""", element, offset)

	def change_attribute_value(self):
		pass

	def smooth_scroll(self):
		pass

	def clear_element_text(self, xpath: str):
		"""Clears the text inside the element.
		"""
		self.wait_for_element_visibility(xpath)
		self.driver.find_element_by_xpath(xpath).clear()

	def send_keys_to_element(self, xpath: str, keys: str):
		"""Fill in the element with the given keys string.
		"""
		self.wait_for_element_visibility(xpath)
		self.driver.find_element_by_xpath(xpath).send_keys(keys)

	def send_keys_using_actions(self, xpath: str, keys: str):
		"""Fill in the element given by the xpath with the given keys string.
		"""
		self.wait_for_element_visibility(xpath)
		self.actions.send_keys(keys)

	def extractor(self, element) -> list:
		parsed_element = None
		if type(element) == list:
			parsed_element = []
			for item in element:
				parsed_element.append(self.extractor(item))
		elif type(element) == dict:
			parsed_element = {}
			for key, value in element.items():
				parsed_element[key] = self.extractor(value)
		elif type(element) == str:
			parsed_element = self.get_text_from_element_by_xpath(element)
		return parsed_element

	def extract(self, element: dict) -> str:
		"""Extracts the text according to the givien dict xpahs.
		
		elements is in the form of {
			'key1' : 'xpath1',
			'key2' : {
				'key3' : 'xpath3',
				'key4' : 'xpath4',
			},
			'key5' : [
				{
					'key6' : 'xpath6',
				},
			]
		}
		"""
		if type(element) != dict:
			return None
		return extractor(element)

	@classmethod
	def get_csv_file_reader(cls, filename: str) -> csv.DictReader:
		"""Opens csv file and returns the csv reader.
		"""
		csv_file = open(filename, newline='')
		return csv.DictReader(csv_file)

	@classmethod
	def get_csv_file_writer(cls, filename: str, fieldnames: list, previlage: str, write_header: bool = True) -> (csv.DictWriter, open):
		"""Opens csv file and writes the header and returns the file writer.
		"""
		csv_file = open(filename, previlage, newline='')
		writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
		if write_header and 'a' not in previlage:
			writer.writeheader()
		return writer, csv_file

	@classmethod
	def convert_csv_to_excel(cls, csv_filename: str, excel_filename: str):
		workbook = Workbook()
		workbook.create_sheet()
		worksheet = workbook.active
		with open(csv_filename, encoding='utf8') as f:
			reader = csv.reader(f)
			for r, row in enumerate(reader):
				for c, col in enumerate(row):
					worksheet.cell(row=r + 1, column=c + 1).value = col
		workbook.save(excel_filename)

	def set_state_parameters(self, state_filename: str = 'state.csv', state_fieldnames: list = []):
		self.state_filename = state_filename
		self.state_fieldnames = state_fieldnames

	def save_state(self, args: list):
		state = {}
		for arg, fieldname in zip(args, self.state_fieldnames):
			state[fieldname] = arg
		writer, file = self.get_csv_file_writer(self.state_filename, self.state_fieldnames, 'w')
		writer.writerow(state)
		file.close()

	def load_state(self):
		reader = self.get_csv_file_reader(self.state_filename)
		self.state_fieldnames = reader.fieldnames
		return next(reader)

	def save_page_content(self, filename: str):
		"""Saves the current page content in a file with the given filename.
		"""
		with open(filename, 'w') as f:
			f.write(self.driver.page_source)

	def initiate(self):
		self.options.initiate()
		self.driver = self.options.driver
		self.wait   = self.options.wait
		self.actions= self.options.actions

	def run(self):
		self.initiate()

	def execute(self, *args):
		"""This function should be overriden to pass custom arguments to the
		run function and also to set up the options.
		Example : self.arg1 = args[0]
		"""
		self.start()
